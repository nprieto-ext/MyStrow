"""
Import d'un patch venu d'un AUTRE logiciel — MyStrow

Deux portes d'entree, sans Qt (ce module est du calcul pur, testable seul) :

  * `parse_qlcplus_workspace()` — un espace de travail QLC+ (.qxw). C'est du XML
    en clair : chaque `<Fixture>` porte Manufacturer / Model / Mode / Universe /
    Address / Channels. Le fichier ne contient PAS la definition des canaux, il
    ne fait que citer la fixture par son nom : c'est la bibliotheque locale qui
    doit la retrouver (voir `match_entry`).

  * `parse_table()` — un tableau CSV / TSV / XLSX. C'est la porte universelle :
    Sunlite, Daslight, Chamsys, Onyx... ne publient pas leur format, mais tous
    savent sortir (ou faire recopier) un tableau « nom / univers / adresse /
    canaux ». C'est aussi la symetrie de `patch_export.export_patch_xlsx`.

Les deux rendent la MEME structure, une liste d'« entrees » :

    {"name", "manufacturer", "model", "mode",
     "universe": 0-based, "address": 1-based, "channels": int|None,
     "group": str|None, "source": str}

`resolve()` complete ensuite chaque entree avec un profil MyStrow et un indice
de confiance. Rien n'est applique ici : l'appelant montre le resultat a
l'utilisateur AVANT de patcher (patch_import_ui.PatchImportDialog).

Deux pieges de numerotation, la raison d'etre de la moitie de ce fichier :
  - QLC+ ecrit l'adresse ET l'univers en base 0 (`<Address>0</Address>` = canal 1) ;
  - un tableau ecrit par un humain les compte a partir de 1.
Tout est ramene ici a la convention MyStrow : univers 0-based, adresse 1-based.
"""

from __future__ import annotations

import csv
import gzip
import json
import os
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
from difflib import get_close_matches
from pathlib import Path

# i18n est du texte pur (pas de Qt) : ce module reste utilisable seul.
from i18n import tr

# Groupes MyStrow (cf. GROUP_BLOCKS dans main_window)
GROUPS = ["face", "lat", "contre", "douche1", "douche2", "douche3",
          "groupe_g", "groupe_h"]

# Types acceptes par le dialogue Patch DMX. Les categories propres a QLC+
# (« Effet », « Laser », « Color Changer »...) n'y existent pas : elles sont
# rabattues, sinon le combo du dialogue les remettrait a « PAR LED » au premier
# clic, en silence.
FIXTURE_TYPES = ["PAR LED", "Moving Head", "Barre LED", "Stroboscope",
                 "Machine a fumee", "Gradateur"]

_TYPE_ALIASES = {
    "effet": "PAR LED", "effect": "PAR LED", "laser": "PAR LED",
    "colorchanger": "PAR LED",
    "dimmer": "Gradateur", "gradateur": "Gradateur",
    "movinghead": "Moving Head", "scanner": "Moving Head", "lyre": "Moving Head",
    "smoke": "Machine a fumee", "fumee": "Machine a fumee",
    "hazer": "Machine a fumee", "fog": "Machine a fumee",
    "machineafumee": "Machine a fumee",
    "strobe": "Stroboscope", "stroboscope": "Stroboscope",
    "ledbar": "Barre LED", "barreled": "Barre LED", "bar": "Barre LED",
    "parled": "PAR LED", "par": "PAR LED",
}

MAX_UNIVERSES = 4          # MyStrow pilote 4 univers Art-Net (0..3)

# Profils de repli quand la fixture reste introuvable. Ce sont les memes listes
# que `artnet_dmx.DMX_PROFILES`, recopiees ici pour ne pas importer le moteur
# DMX (donc Qt, donc l'application entiere) depuis un simple parseur.
_GENERIC = {
    1: ["Dim"],
    2: ["Dim", "Strobe"],
    3: ["R", "G", "B"],
    4: ["R", "G", "B", "Dim"],
    5: ["R", "G", "B", "Dim", "Strobe"],
    6: ["R", "G", "B", "W", "Dim", "Strobe"],
}
_GENERIC_MOVING = ["Pan", "Tilt", "Dim", "ColorWheel", "Gobo1", "Shutter",
                   "Speed", "Mode"]

# Trois noms de canaux de la base QLC+ n'existent pas dans `artnet_dmx` : ils
# sortiraient 0 en dur, donc un canal mort a l'arrivee (l'ambre de 1080 modes,
# la rotation de gobo de 56). Ils sont recales ici, a la lecture.
_CH_ALIASES = {"Effect": "Effects", "A": "Ambre", "GoboRot": "Gobo1Rot"}

# Lettre de groupe telle que l'ecrit `patch_export` (colonne « Groupe » du
# classeur Excel). C'est ce qui rend le trajet export -> reimport possible.
_GROUP_LETTERS = {"a": "face", "b": "lat", "c": "contre", "d": "douche1",
                  "e": "douche2", "f": "douche3", "g": "groupe_g",
                  "h": "groupe_h"}


def _known_channels() -> set:
    """Types de canaux que le moteur DMX sait piloter. Source unique :
    `artnet_dmx.CHANNEL_TYPES` (module sans Qt, importable ici)."""
    try:
        from artnet_dmx import CHANNEL_TYPES
        return set(CHANNEL_TYPES)
    except Exception:
        return set()


# ──────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ──────────────────────────────────────────────────────────────────────────────

def _norm(s) -> str:
    """Cle de comparaison : sans accents, sans ponctuation, sans casse.

    « Mac Aura XB » et « MAC-Aura_XB » doivent tomber sur la meme cle, sinon la
    correspondance exacte rate sur une difference de tiret.
    """
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _res(name: str) -> str:
    """Chemin d'une ressource embarquee (dossier du script ou _MEIPASS)."""
    here = os.path.dirname(os.path.abspath(__file__))
    base = getattr(sys, "_MEIPASS", here)
    p = os.path.join(base, name)
    return p if os.path.exists(p) else os.path.join(here, name)


def _load_json(path):
    try:
        if str(path).endswith(".gz"):
            with gzip.open(path, "rb") as f:
                return json.loads(f.read().decode("utf-8"))
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _int(v, default=None):
    try:
        return int(str(v).strip())
    except Exception:
        return default


def type_from_profile(profile) -> str:
    """Devine le type a partir des canaux — dernier recours."""
    s = set(profile or [])
    if {"Pan", "Tilt"} & s:
        return "Moving Head"
    if {"Smoke", "Fan"} & s:
        return "Machine a fumee"
    if s and s <= {"Dim", "Dim2", "Unused"}:
        return "Gradateur"
    if "Strobe" in s and not ({"R", "G", "B"} & s):
        return "Stroboscope"
    return "PAR LED"


def normalize_type(raw, profile=None) -> str:
    """Ramene un type venu d'ailleurs a l'un des `FIXTURE_TYPES`."""
    key = (raw or "").strip()
    if key in FIXTURE_TYPES:
        return key
    alias = _TYPE_ALIASES.get(_norm(key))
    if alias:
        return alias
    return type_from_profile(profile or [])


def guess_group(*texts) -> str | None:
    """Devine le groupe MyStrow d'apres un libelle (nom de fixture, de groupe).

    Aucun autre logiciel n'a la notion de face / douche / contre : c'est propre
    a MyStrow. Le seul indice disponible est le texte que le regisseur a tape.
    """
    blob = " ".join(_norm(t) for t in texts if t)
    if not blob:
        return None
    for kw, grp in (
        ("contre", "contre"), ("backlight", "contre"), ("gegen", "contre"),
        ("back", "contre"),
        ("douche1", "douche1"), ("douche2", "douche2"), ("douche3", "douche3"),
        ("dch1", "douche1"), ("dch2", "douche2"), ("dch3", "douche3"),
        ("douche", "douche1"), ("top", "douche1"), ("down", "douche1"),
        ("laterale", "lat"), ("lateral", "lat"), ("side", "lat"),
        ("cote", "lat"), ("lat", "lat"),
        ("frontal", "face"), ("front", "face"), ("face", "face"),
    ):
        if kw in blob:
            return grp
    return None


def group_from_cell(raw) -> str | None:
    """Lit une cellule « Groupe » : nom MyStrow, lettre A–H, ou texte libre."""
    s = (raw or "").strip()
    if not s:
        return None
    if s in GROUPS:
        return s
    if len(s) == 1 and s.lower() in _GROUP_LETTERS:
        return _GROUP_LETTERS[s.lower()]
    return guess_group(s)


def parse_profile_cell(raw):
    """Lit une liste de canaux ecrite dans une cellule.

    C'est la colonne « Profil DMX » du classeur exporte par `patch_export`
    (« R · G · B · Dim »), et c'est aussi la seule facon, pour un logiciel dont
    on ne sait pas lire le fichier, de decrire ses appareils au canal pres :
    l'utilisateur tape la liste dans son tableau.
    """
    s = (raw or "").strip()
    if not s:
        return []
    parts = [p.strip() for p in re.split(r"[·,;|/]+|\s{2,}", s) if p.strip()]
    if len(parts) <= 1:
        parts = [p for p in s.split() if p]
    known = _known_channels()
    # Correspondance insensible a la casse : un tableau tape a la main ecrit
    # « dim », « STROBE », « Pan ». Un nom inconnu devient « Unused » — le canal
    # garde sa place, donc l'appareil suivant reste a la bonne adresse.
    lookup = {_norm(k): k for k in known}
    out = []
    for p in parts:
        p = _CH_ALIASES.get(p, p)
        out.append(lookup.get(_norm(p), "Unused"))
    return out


# ──────────────────────────────────────────────────────────────────────────────
# Bibliotheque locale — index de correspondance
# ──────────────────────────────────────────────────────────────────────────────

def _iter_library_records():
    """Toutes les fixtures connues, une entree par MODE.

    Meme assemblage que le dialogue « Bibliotheque de fixtures » du patch, dans
    le meme ordre de priorite : QLC+ d'abord (c'est la source d'un .qxw, donc
    celle qui donnera la correspondance exacte), puis les profils integres,
    puis ce que l'utilisateur a cree ou telecharge.
    """
    # 1) Base QLC+ — {manufacturer, model, fixture_type, modes:[{name, channels}]}
    for fx in (_load_json(_res("fixtures_qlcplus.json")) or []):
        mfr = fx.get("manufacturer", "")
        model = fx.get("model", "")
        ftype = fx.get("fixture_type", "")
        for m in fx.get("modes", []):
            chans = list(m.get("channels") or [])
            if not chans:
                continue
            yield {"mfr": mfr, "model": model, "mode": m.get("name", ""),
                   "profile": chans,
                   "fixture_type": normalize_type(ftype, chans),
                   "source": "qlcplus"}

    # 2) Profils integres — un seul mode, le nom porte deja le nombre de canaux
    try:
        from builtin_fixtures import BUILTIN_FIXTURES
    except Exception:
        BUILTIN_FIXTURES = []
    for fx in BUILTIN_FIXTURES:
        prof = list(fx.get("profile") or [])
        if not prof:
            continue
        yield {"mfr": fx.get("manufacturer", ""), "model": fx.get("name", ""),
               "mode": "", "profile": prof,
               "fixture_type": normalize_type(fx.get("fixture_type"), prof),
               "source": "builtin"}

    # 3) Fixtures de l'utilisateur + packs telecharges + bundles embarques
    sources = []
    user_file = Path.home() / ".mystrow_fixtures.json"
    if user_file.exists():
        sources.append(_load_json(str(user_file)))
    sources.append(_load_json(_res("fixtures_bundle_custom.json.gz")))
    sources.append(_load_json(_res("fixtures_bundle.json.gz")))
    for data in sources:
        for fx in (data or []):
            if not isinstance(fx, dict):
                continue
            mfr = fx.get("manufacturer", "")
            model = fx.get("name", "")
            for m in (fx.get("modes") or []):
                prof = list(m.get("profile") or [])
                if prof:
                    yield {"mfr": mfr, "model": model,
                           "mode": m.get("name", ""), "profile": prof,
                           "fixture_type": normalize_type(fx.get("fixture_type"), prof),
                           "source": fx.get("source", "user")}
            prof = list(fx.get("profile") or [])
            if prof:
                yield {"mfr": mfr, "model": model, "mode": "", "profile": prof,
                       "fixture_type": normalize_type(fx.get("fixture_type"), prof),
                       "source": fx.get("source", "user")}


_INDEX_CACHE = None


def build_index(force=False) -> dict:
    """Index de recherche : par (marque, modele), par modele seul, et la liste
    des modeles normalises pour la recherche approchante."""
    global _INDEX_CACHE
    if _INDEX_CACHE is not None and not force:
        return _INDEX_CACHE
    by_pair, by_model = {}, {}
    for rec in _iter_library_records():
        nm, nmo = _norm(rec["mfr"]), _norm(rec["model"])
        if not nmo:
            continue
        rec["profile"] = [_CH_ALIASES.get(c, c) for c in rec["profile"]]
        label = " ".join(x for x in (rec["mfr"], rec["model"]) if x)
        if rec["mode"]:
            label += f" — {rec['mode']}"
        rec["_label"] = label
        by_pair.setdefault((nm, nmo), []).append(rec)
        by_model.setdefault(nmo, []).append(rec)
    _INDEX_CACHE = {"by_pair": by_pair, "by_model": by_model,
                    "models": list(by_model.keys())}
    return _INDEX_CACHE


def _pick(cands, channels, mode):
    """Parmi des candidats, celui dont le mode colle — sinon le nombre de canaux."""
    if not cands:
        return None, None
    nmode = _norm(mode)
    if nmode:
        for c in cands:
            if _norm(c["mode"]) == nmode:
                return c, "exact"
    if channels:
        same = [c for c in cands if len(c["profile"]) == channels]
        if same:
            if nmode:
                near = get_close_matches(nmode, [_norm(c["mode"]) for c in same],
                                         n=1, cutoff=0.6)
                if near:
                    for c in same:
                        if _norm(c["mode"]) == near[0]:
                            return c, "good"
            return same[0], "good"
        return None, None
    return cands[0], "good"


def generic_profile(channels, fixture_type=None) -> list:
    """Profil de repli quand la fixture est introuvable : le patch reste juste
    (bon nombre de canaux, bonne adresse suivante), seul le detail est a revoir."""
    n = max(1, int(channels or 1))
    if fixture_type == "Moving Head":
        base = list(_GENERIC_MOVING)
    elif n in _GENERIC:
        return list(_GENERIC[n])
    else:
        base = list(_GENERIC[6])
    if len(base) >= n:
        return base[:n]
    return base + ["Unused"] * (n - len(base))


def match_entry(entry, index=None) -> dict:
    """Cherche le profil MyStrow d'une entree de patch.

    Quatre niveaux, du plus sur au plus approximatif :
      exact   — marque + modele + mode identiques ;
      good    — marque + modele, mode retrouve par le nombre de canaux ;
      approx  — modele seul, ou orthographe voisine ;
      generic — rien trouve : profil deduit du nombre de canaux.
    """
    idx = index or build_index()
    mfr = entry.get("manufacturer", "")
    model = entry.get("model", "")
    mode = entry.get("mode", "")
    chans = entry.get("channels")
    nm, nmo = _norm(mfr), _norm(model)

    rec, conf = _pick(idx["by_pair"].get((nm, nmo), []), chans, mode)
    if rec is None:
        rec, conf = _pick(idx["by_model"].get(nmo, []), chans, mode)
        if rec is not None:
            conf = "good" if conf == "exact" else "approx"
    if rec is None and nmo:
        near = get_close_matches(nmo, idx["models"], n=1, cutoff=0.86)
        if near:
            rec, _c = _pick(idx["by_model"][near[0]], chans, mode)
            if rec is not None:
                conf = "approx"

    if rec is None and chans == 1:
        # Un appareil d'un seul canal EST un gradateur : rien a deviner, rien a
        # faire relire. C'est le cas des blocs de puissance, tres nombreux dans
        # un patch importe (QLC+ les ecrit en « Generic / Generic »).
        return {"profile": ["Dim"], "fixture_type": "Gradateur",
                "matched": tr("pimp_i_dimmer_1ch"), "confidence": "good"}

    if rec is None:
        ftype = normalize_type(entry.get("fixture_type"), None)
        prof = generic_profile(chans, ftype)
        return {"profile": prof,
                "fixture_type": normalize_type(entry.get("fixture_type"), prof),
                "matched": "", "confidence": "generic"}

    prof = list(rec["profile"])
    # Le fichier source fait foi sur l'ENCOMBREMENT : si le profil trouve n'a pas
    # le meme nombre de canaux, la fixture suivante serait decalee. On aligne, et
    # on retrograde la confiance — c'est exactement le cas a faire relire.
    if chans and len(prof) != chans:
        prof = (prof + ["Unused"] * chans)[:chans]
        conf = "approx"
    return {"profile": prof, "fixture_type": rec["fixture_type"],
            "matched": rec["_label"], "confidence": conf or "approx"}


# ──────────────────────────────────────────────────────────────────────────────
# QLC+ (.qxw)
# ──────────────────────────────────────────────────────────────────────────────

def _strip_ns(data: bytes) -> bytes:
    """QLC+ ecrit un espace de noms sur la racine ; il polluerait chaque tag."""
    return re.sub(rb'\sxmlns(:\w+)?="[^"]*"', b"", data)


def _txt(el, tag, default=""):
    child = el.find(tag)
    if child is not None and child.text:
        return child.text.strip()
    return default


def parse_qlcplus_workspace(path):
    """Lit un .qxw et rend (entrees, avertissements).

    Rappel du format : `<Address>` et `<Universe>` sont en base 0.
    """
    warnings = []
    with open(path, "rb") as f:
        raw = f.read()
    root = ET.fromstring(_strip_ns(raw))
    engine = root.find("Engine")
    if engine is None:
        engine = root

    # Groupes QLC+ : le seul indice pour retrouver un groupe MyStrow. Le tag
    # « Head » a change de forme entre les versions (attribut ou texte) — on
    # ramasse le premier entier qu'on trouve, quelle que soit la forme.
    group_of = {}
    for grp in engine.iter("FixtureGroup"):
        gname = _txt(grp, "Name")
        if not gname:
            continue
        for head in grp.iter():
            if "head" not in head.tag.lower():
                continue
            ids = []
            for key in ("Fixture", "fixture", "ID", "id"):
                v = _int(head.get(key))
                if v is not None:
                    ids.append(v)
            if not ids and head.text:
                m = re.match(r"\s*(\d+)", head.text)
                if m:
                    ids.append(int(m.group(1)))
            for fid in ids:
                group_of.setdefault(fid, gname)

    entries = []
    for fx in engine.iter("Fixture"):
        if fx.find("Model") is None and fx.find("Channels") is None:
            continue          # <Fixture> d'un autre contexte (groupe, scene)
        fid = _int(_txt(fx, "ID"))
        mfr = _txt(fx, "Manufacturer")
        model = _txt(fx, "Model")
        mode = _txt(fx, "Mode")
        name = _txt(fx, "Name") or model
        chans = _int(_txt(fx, "Channels"))
        uni = _int(_txt(fx, "Universe"), 0) or 0
        addr = _int(_txt(fx, "Address"), 0) or 0
        gname = group_of.get(fid, "")

        # Gradateur generique QLC+ : une fixture « Generic/Generic » de N canaux,
        # ce sont N gradateurs INDEPENDANTS. La garder d'un bloc les ferait tous
        # bouger ensemble — on la deplie.
        if _norm(mfr) == "generic" and _norm(model) == "generic" and (chans or 0) > 1:
            for k in range(chans):
                entries.append({
                    "name": f"{name} {k + 1}", "manufacturer": mfr,
                    "model": model, "mode": mode, "universe": uni,
                    "address": addr + k + 1, "channels": 1,
                    "fixture_type": "Gradateur",
                    "group": guess_group(name, gname), "source": "qlcplus",
                })
            warnings.append(tr("pimp_i_dimmer_split", name=name, n=chans))
            continue

        entries.append({
            "name": name, "manufacturer": mfr, "model": model, "mode": mode,
            "universe": uni, "address": addr + 1, "channels": chans,
            "fixture_type": "", "group": guess_group(name, gname),
            "source": "qlcplus",
        })

    if not entries:
        warnings.append(tr("pimp_i_no_fixture_qlc"))
    return entries, warnings


# ──────────────────────────────────────────────────────────────────────────────
# Tableau (CSV / TSV / XLSX)
# ──────────────────────────────────────────────────────────────────────────────

_COLS = {
    "name":         ("nom", "name", "fixture", "appareil", "libelle", "label",
                     "projecteur", "instrument"),
    "manufacturer": ("marque", "fabricant", "manufacturer", "constructeur",
                     "brand", "hersteller"),
    "model":        ("modele", "model", "produit", "reference", "ref"),
    "mode":         ("mode", "modedmx", "dmxmode"),
    "universe":     ("univers", "universe", "uni", "u", "output", "sortie"),
    "address":      ("adresse", "address", "addr", "dmx", "adressedmx",
                     "adressedepart", "startaddress", "patch", "start"),
    "channels":     ("canaux", "channels", "nbcanaux", "nbch", "ch",
                     "numchannels", "footprint", "nbcanal"),
    "group":        ("groupe", "group", "zone", "categorie"),
    "type":         ("type", "typefixture", "fixturetype"),
    "profile":      ("profil", "profile", "profildmx", "dmxprofile",
                     "canauxdmx", "profilcanaux"),
}


def _map_columns(header):
    """Associe chaque colonne du fichier a un champ connu. Rend {champ: index}."""
    out = {}
    for i, cell in enumerate(header):
        n = _norm(cell)
        if not n:
            continue
        hit = None
        for field, aliases in _COLS.items():
            if field not in out and n in aliases:
                hit = field
                break
        if hit is None:
            # Deuxieme passe, plus permissive : « adresse dmx de depart »
            for field, aliases in _COLS.items():
                if field in out:
                    continue
                if any(a in n for a in aliases if len(a) > 3):
                    hit = field
                    break
        if hit:
            out[hit] = i
    return out


def _parse_address(raw):
    """Rend (univers|None, adresse|None).

    Accepte « 45 », « 1.045 », « 2/12 », « U2 CH 45 » — les ecritures qu'on
    trouve dans un export MA, Chamsys ou une feuille tapee a la main. Sur la
    forme combinee, l'univers est compte a partir de 1 (convention humaine).
    """
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    m = re.search(r"(\d+)\s*[./:]\s*(\d+)", s)
    if m:
        return max(0, int(m.group(1)) - 1), int(m.group(2))
    m = re.search(r"(\d+)", s)
    return None, (int(m.group(1)) if m else None)


def _read_rows(path):
    """Rend la liste des lignes (listes de chaines) d'un CSV/TSV/XLSX."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                text = f.read()
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(tr("pimp_i_encoding"))
    sample = "\n".join(text.splitlines()[:5])
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except Exception:
        counts = {d: sample.count(d) for d in ";,\t|"}
        delim = max(counts, key=counts.get) if any(counts.values()) else ";"
    return list(csv.reader(text.splitlines(), delimiter=delim))


def parse_table(path):
    """Lit un tableau de patch et rend (entrees, avertissements).

    Convention humaine : univers et adresse comptes a partir de 1. L'univers
    ressort en base 0 (convention MyStrow).
    """
    warnings = []
    rows = [r for r in _read_rows(path) if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(tr("pimp_i_empty_file"))

    # La ligne d'en-tete n'est pas toujours la premiere (titre, date...) :
    # on prend la premiere ligne ou l'on reconnait au moins deux colonnes.
    head_i, cols = None, {}
    for i, row in enumerate(rows[:10]):
        c = _map_columns(row)
        if len(c) >= 2 and ("address" in c or "channels" in c):
            head_i, cols = i, c
            break
    if head_i is None:
        raise ValueError(tr("pimp_i_no_header"))

    entries = []
    for line_no, row in enumerate(rows[head_i + 1:], start=head_i + 2):
        def cell(field):
            i = cols.get(field)
            return str(row[i]).strip() if i is not None and i < len(row) else ""

        uni_c, addr = _parse_address(cell("address"))
        if addr is None:
            continue                       # ligne de commentaire ou de total
        uni_col = _int(cell("universe"))
        if uni_col is not None:
            uni = max(0, uni_col - 1)      # tableau humain : univers 1 = index 0
        elif uni_c is not None:
            uni = uni_c
        else:
            uni = 0
        chans = _int(cell("channels"))
        name = cell("name") or cell("model") or f"Appareil {len(entries) + 1}"
        # Colonne « Profil DMX » : elle decrit l'appareil canal par canal, donc
        # il n'y a plus rien a deviner. C'est ce qui rend le trajet
        # patch_export -> reimport exact, et ce qui sauve les logiciels dont on
        # ne sait pas lire le fichier (Sunlite : l'utilisateur tape la liste).
        prof = parse_profile_cell(cell("profile"))
        if prof and chans and len(prof) != chans:
            warnings.append(tr("pimp_i_ch_mismatch", line=line_no, name=name,
                               got=len(prof), want=chans))
            chans = len(prof)
        elif prof:
            chans = len(prof)
        elif chans is None:
            warnings.append(tr("pimp_i_no_channels", line=line_no, name=name))
        entries.append({
            "name": name,
            "manufacturer": cell("manufacturer"),
            "model": cell("model") or name,
            "mode": cell("mode"),
            "universe": uni, "address": addr, "channels": chans,
            "fixture_type": cell("type"),
            "group": group_from_cell(cell("group")) or guess_group(name),
            "profile": prof,
            "source": "table",
        })

    if not entries:
        warnings.append(tr("pimp_i_no_row"))
    return entries, warnings


# ──────────────────────────────────────────────────────────────────────────────
# Resolution + controles
# ──────────────────────────────────────────────────────────────────────────────

def resolve(entries):
    """Complete les entrees : profil, type, groupe, et les problemes a montrer.

    Chaque ligne ressort avec `issues`, une liste de messages. Une ligne qui a
    un probleme BLOQUANT (`blocking`) est decochee d'office dans le dialogue :
    l'importer casserait le patch (hors univers, depassement du canal 512).
    """
    idx = build_index()
    rows = []
    for e in entries:
        given = e.get("profile") or []
        if given:
            # Le fichier decrit lui-meme les canaux : plus rien a chercher.
            m = {"profile": list(given),
                 "fixture_type": type_from_profile(given),
                 "matched": tr("pimp_i_profile_from_file"),
                 "confidence": "exact"}
        else:
            m = match_entry(e, idx)
        chans = e.get("channels") or len(m["profile"])
        row = dict(e)
        row.update({
            "profile": m["profile"],
            "fixture_type": (normalize_type(e.get("fixture_type"), m["profile"])
                             if e.get("fixture_type") else m["fixture_type"]),
            "matched": m["matched"],
            "confidence": m["confidence"],
            "channels": chans,
            "group": e.get("group") or "face",
            "issues": [], "blocking": False, "include": True,
        })
        if row["universe"] >= MAX_UNIVERSES:
            row["issues"].append(
                tr("pimp_i_universe", u=row["universe"] + 1, max=MAX_UNIVERSES))
            row["blocking"] = True
        if row["address"] < 1 or row["address"] + chans - 1 > 512:
            row["issues"].append(
                tr("pimp_i_overflow", a=row["address"], n=chans))
            row["blocking"] = True
        if m["confidence"] == "generic":
            row["issues"].append(tr("pimp_i_unknown"))
        elif m["confidence"] == "approx":
            row["issues"].append(tr(
                "pimp_i_approx", m=m["matched"] or tr("pimp_i_adjusted")))
        row["include"] = not row["blocking"]
        rows.append(row)

    # Chevauchements d'adresses — le vrai piege d'un patch importe d'ailleurs,
    # ou une fixture a pu etre patchee a la main par-dessus une autre.
    occupied = {}
    for i, r in enumerate(rows):
        for ch in range(r["address"], r["address"] + r["channels"]):
            key = (r["universe"], ch)
            other = occupied.get(key)
            if other is None:
                occupied[key] = i
            elif other != i:
                msg = tr("pimp_i_overlap", name=rows[other]["name"])
                if msg not in r["issues"]:
                    r["issues"].append(msg)
    return rows


def summarize(rows) -> dict:
    """Compte par niveau de confiance — le resume affiche en tete du dialogue."""
    out = {"total": len(rows), "exact": 0, "good": 0, "approx": 0,
           "generic": 0, "blocking": 0}
    for r in rows:
        out[r["confidence"]] = out.get(r["confidence"], 0) + 1
        if r["blocking"]:
            out["blocking"] += 1
    return out
